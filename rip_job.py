from pyats.easypy import run
import os,atexit,re
import zipfile
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font,PatternFill
#from log_list import command
global report_dir
global router_name,str1,l

router_name = ['n9k1#','n9k2#','n9k3#','NEXUS9k_4#']
str1 = ['show version', 'show access']
l = []

def report_html():
    global report_dir
    extract_path = "/home/tcs/POC2/PYATS/RIP_Dummy/result/"
    with zipfile.ZipFile(report_dir, 'r') as zip_ref:
        zip_ref.extractall(extract_path)

    html_op_file = "/home/tcs/POC2/PYATS/RIP_Dummy/"

    res_report = os.path.join(extract_path,"rip_job.report")
    with open(res_report,'r') as res:
        lines = res.readlines()

    data = []
    for z in lines:
        parts = z.strip().split(':',1)
        if len(parts) == 2:
            if parts[0].startswith('Task') and len(parts[1].split())>1:
                a = parts[1].split()
                key = parts[0].strip()
                value = a[0].strip()
                status = a[1].strip()
                data.append([key,value,status])
            else:
                if parts[0].startswith('Task') and len(parts[1].split())==1:
                    pass
                else:
                    key,value = parts[0].strip(),parts[1].strip()
                    status = ''
                    data.append([key,value,status])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Attribute','Output','Status'])
    bold_font = Font(bold=True)
    row_num = 1
    for cell in ws[row_num]:
         cell.font = bold_font

    for key in data:
         ws.append([key[0],key[1],key[2]])

    green_fill = PatternFill(start_color = '00FF00', end_color = '00FF00',fill_type = 'solid')
    red_fill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
    yellow_fill = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')
    default_fill = PatternFill(start_color = 'FFFFFF', end_color = 'FFFFFF', fill_type = 'solid')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value == "PASSED":
                cell.fill = green_fill
            elif cell.value == "FAILED":
                cell.fill = red_fill
            elif cell.value == 'ERRORED':
                cell.fill = yellow_fill
            else:
                cell.fill = default_fill

    op_file = os.path.join(html_op_file,'pyxl_report.xlsx')
    wb.save(op_file)

    html_path = os.path.join(html_op_file,'final_report.html')

    df = pd.read_excel(op_file)
    html_table = df.to_html()
    with open(html_path,'w') as html_file:
         html_file.write(html_table)

    print('Link to open HTML report result :',html_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Command','Output'])

    pattern = r'TaskLog\.Task-\d+'
    dir_path = '/home/tcs/POC2/PYATS/RIP_Dummy/result/'

    file_list = os.listdir(dir_path)
    file_list.sort()

    matching_file = []

    for f in file_list:
        if re.match(pattern,f):
            matching_file.append(f)

    file_content=[]
    for k in matching_file:
        pt = os.path.join(dir_path,k)
        cont = open(pt).read()
        file_content.append(cont)

   # command(str1,router_name,file_content,ws,l)

    wb.save('output_log.xlsx')

    final_html = open('final.html','w')
    ht = """<html><head><title>Log Result</title></head><body>"""
    for d in l:
        ht = ht + """<p>{}</p><pre>{}</pre>""".format(d[0],d[1])
    ht = ht + '</body></html>'
    final_html.write(ht)
    final_html.close()

    f_path = os.path.abspath('final.html')
    print('Link to view Log results :',f_path)

def main(runtime):
    global report_dir
    atexit.register(report_html)
    report_dir = runtime.archive
    script_file = ['testcase_rip_all.py']
    for x in script_file:
        run(x, datafile = 'rip_datafile.yaml',pdb=True)
